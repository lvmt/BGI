#!/usr/bin/python3
# -*- coding:utf-8 -*-
# @Author lvmengting
# @Time   2020/12/31 15:44
# @Email  13554221497@163.com
# @File   mutation_in_bed.py


"""
purpose:
统计文献中的变异位点是否在BGI 淋巴瘤的bed区间内
"""


import re
from openpyxl import Workbook
from openpyxl import load_workbook
from collections import defaultdict


def is_overlap(_chr1, start1, end1, _chr2, start2, end2, cutoff):
    """
    Args:
        _chr1: 变异的染色体位置
        start1:
        end1:
        _chr2:  bed库文件的染色体位置
        start2:
        end2:
        cutoff: overlap / len(end1 - start1)

    Returns:

    """
    if end1 < start2 or end2 < start1:
        return False
    elif max(start1, start2) <= min(end1, end2):
        overlap = min(end1, end2) - max(start1, start2) + 1
        percent = float(overlap) / float(end1 - start1 + 1)
        if percent >= cutoff:
            return True
        return False


def get_bed_dict(bedfile):
    """
    根据问库文件bed, 构建查询字典
    dict
    {
        chr1: [(start, end), (start, end)...],
        chr2: [(start, end), (start, end)...],
        ..
    }
    """
    bed_dict = defaultdict(list)
    with open(bedfile, 'r') as fr:
        for line in fr:
            linelist = line.strip().split('\t')
            bed_dict[linelist[0]].append((int(linelist[1]), int(linelist[2])))
    return bed_dict


if __name__ == '__main__':
    import sys
    infile = sys.argv[1]
    outfile = sys.argv[2]
    bedfile = sys.argv[3]

    bed_dict = get_bed_dict(bedfile)

    wb = Workbook()
    out_sheet = wb.create_sheet('变异结果', index=0)

    wr = load_workbook(infile)
    sheet_names = wr.sheetnames
    read_sheet = wr[sheet_names[0]]
    n = 1
    for row in read_sheet.rows:
        values = [cell.value for cell in row]
        tag = 'not in bed'
        if n == 1:
            out_sheet.append(values)
        elif n == 2:
            values.insert(1, 'mutation in bed')
            out_sheet.append(values)
        else:
            # print(values[0])
            info = values[0].split('_')
            _chr1 = info[0]
            start1 = int(info[1])
            end1 = int(info[2])

            sub_bed_dict = bed_dict[_chr1]
            for (start2, end2) in sub_bed_dict:
                if is_overlap(_chr1, start1, end1, _chr1, start2, end2, 1):
                    tag = 'in bed'
                    break
            values.insert(1, tag)
            out_sheet.append(values)
        n += 1
    wb.save(outfile)


